import os
import sys
import requests
import json
import re
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env'))

url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")

if not url or not key:
    print("Error: SUPABASE_URL or SUPABASE_KEY not found in .env")
    sys.exit(1)

REST_URL = f"{url}/rest/v1"
HEADERS = {
    "apikey": key,
    "Authorization": f"Bearer {key}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

def generate_slug(name):
    slug = name.lower()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'\s+', '-', slug)
    return slug

def post_data(table, data):
    resp = requests.post(f"{REST_URL}/{table}", headers=HEADERS, json=data)
    return resp

def import_clinics(ws):
    print("\n--- Importing Clinics ---")
    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        name = row[header_map.get('Clinic Name')]
        if not name: continue
        
        slug = generate_slug(name)
        print(f"Processing {name}...")

        # Safe extraction
        def get_val(col_name):
            idx = header_map.get(col_name)
            return row[idx] if idx is not None else None

        business_data = {
            "slug": slug,
            "name": name,
            "category": "clinic",
            "description": f"{name} - {get_val('Type') or 'Clinic'}",
            "address": get_val('Address'),
            "area": get_val('Area'),
            "phone": str(get_val('Phone') or ''),
            "gmap_link": get_val('Google Maps Link'),
            "working_hours_raw": str(get_val('Working Hours') or ''),
            "whatsapp": str(get_val('Phone') or '') if str(get_val('WhatsApp (Yes/No)')).lower() in ['yes', 'y'] else "",
            "source": get_val('Source Directory')
        }

        resp = post_data('businesses', business_data)
        if resp.status_code == 201:
            bid = resp.json()[0]['id']
            details = {
                "business_id": bid,
                "doctor_name": get_val('Doctor Name(s)'),
                "specialization": get_val('Type'),
                "services_raw": get_val('Services Mentioned')
            }
            post_data('clinic_details', details)
            print("  -> Success")
        elif resp.status_code == 409:
            print("  -> Skipped (Exists)")
        else:
            print(f"  -> Failed: {resp.text}")

def import_shoes(ws):
    print("\n--- Importing Shoe Stores ---")
    # NO HEADERS in 'shoes' sheet based on inspection. Data starts row 1.
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    for row in rows:
        name = row[0]
        if not name: continue
        
        slug = generate_slug(name)
        print(f"Processing {name}...")

        # Mapping based on user provided data structure:
        # 0: Name, 1: Type, 2: Address, 3: Area, 4: Phone, 5: Map, 6: ?, 7: WA?, 8: Hours, 9: Source
        
        # Safety check for row length
        if len(row) < 6: 
             print(f"Skipping row {name} - insufficient data")
             continue

        business_data = {
            "slug": slug,
            "name": name,
            "category": "shoe_store",
            "description": f"{name} - {row[1] if len(row) > 1 else 'Shoe Store'}",
            "address": row[2] if len(row) > 2 else '',
            "area": row[3] if len(row) > 3 else '',
            "phone": str(row[4] if len(row) > 4 else ''),
            "gmap_link": row[5] if len(row) > 5 else '',
            "working_hours_raw": str(row[8]) if len(row) > 8 else '',
            "source": row[9] if len(row) > 9 else 'Google Maps'
        }
        
        # WhatsApp logic seems to be around col 7 (Yes/No)
        if len(row) > 7 and str(row[7]).lower() in ['yes', 'y']:
             business_data['whatsapp'] = business_data['phone']

        resp = post_data('businesses', business_data)
        if resp.status_code == 201:
            bid = resp.json()[0]['id']
            details = {
                "business_id": bid,
                "store_type": row[1], # Raw text import now allowed by migration_v2
                "primary_brand_focus": "" # Not in excel
            }
            post_data('shoe_store_details', details)
            print("  -> Success")
        elif resp.status_code == 409:
            print("  -> Skipped (Exists)")
        else:
            print(f"  -> Failed: {resp.text}")

def import_leather(ws):
    print("\n--- Importing Leather Companies ---")
    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        name = row[header_map.get('Company Name')]
        if not name: continue
        
        slug = generate_slug(name)
        print(f"Processing {name}...")

        def get_val(col_name):
            idx = header_map.get(col_name)
            return row[idx] if idx is not None else None

        business_data = {
            "slug": slug,
            "name": name,
            "category": "leather_company",
            "description": f"{name} - {get_val('Business Type')}",
            "address": get_val('Address'),
            "area": get_val('Area'),
            "phone": str(get_val('Phone') or ''),
            "email": get_val('Email'),
            "gmap_link": get_val('Google Maps Link'),
            "source": get_val('Source Directory')
        }

        resp = post_data('businesses', business_data)
        if resp.status_code == 201:
            bid = resp.json()[0]['id']
            
            export_val = str(get_val('Export Oriented')).lower()
            is_exporter = 'yes' in export_val
            
            details = {
                "business_id": bid,
                "export_status": is_exporter,
                "product_type_raw": get_val('Product Type'),
                "certifications": [get_val('Certifications')] if get_val('Certifications') and get_val('Certifications') != 'No' else []
            }
            post_data('leather_company_details', details)
            print("  -> Success")
        elif resp.status_code == 409:
            print("  -> Skipped (Exists)")
        else:
            print(f"  -> Failed: {resp.text}")

def main():
    try:
        wb = load_workbook('web_site_clients.xlsx')
        
        if 'clinic data' in wb.sheetnames:
            import_clinics(wb['clinic data'])
            
        if 'shoes' in wb.sheetnames:
            import_shoes(wb['shoes'])
            
        if 'leather good' in wb.sheetnames:
            import_leather(wb['leather good'])
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
